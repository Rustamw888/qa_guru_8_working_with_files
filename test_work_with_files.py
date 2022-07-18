import csv
import os
import zipfile

from PyPDF2 import PdfReader
from openpyxl import load_workbook


def test_zip_files():
    fantasy_zip = zipfile.ZipFile('resources/archive.zip', 'w')
    for folder, subfolders, files in os.walk('resources'):

        for file in files:
            if file.endswith('.pdf') or file.endswith('.xlsx') or file.endswith('.csv'):
                fantasy_zip.write(os.path.join(folder, file), os.path.relpath
                (os.path.join(folder,file), 'resources'), compress_type = zipfile.ZIP_DEFLATED)
    fantasy_zip.close()


def test_unzip_files():
    fantasy_zip = zipfile.ZipFile('resources/archive.zip')
    fantasy_zip.extractall('unzipped/')
    fantasy_zip.close()


def test_read_pdf():
    reader = PdfReader('unzipped/docs-pytest-org-en-latest.pdf')
    page = reader.pages[0]
    text = page.extract_text()
    assert '2022' in text


def test_read_xlsx():
    workbook = load_workbook('unzipped/sample1.xlsx')
    sheet = workbook.active
    name = sheet.cell(row=3, column=3).value
    assert 'Ashish' == name


def test_read_csv_():
    with open('unzipped/username.csv') as f:
        reader = csv.reader(f)
        headers = next(reader)
    assert 'Username' in str(headers)